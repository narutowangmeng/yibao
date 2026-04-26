from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
MATERIALS_DIR = ROOT.parent / "材料" / "目录规则示例"
OUTPUT_FILE = ROOT / "src" / "data" / "insuranceKnowledgeBase.json"


def sheet_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    records: list[dict[str, Any]] = []
    for row in rows[1:]:
      record: dict[str, Any] = {}
      empty = True
      for index, header in enumerate(headers):
          key = header or f"column_{index}"
          value = row[index] if index < len(row) else None
          if value not in (None, ""):
              empty = False
          record[key] = value
      if not empty:
          records.append(record)
    return records


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def build_text_entries(records: list[dict[str, Any]], kind: str) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    for index, record in enumerate(records, start=1):
        entry = {
            "id": f"{kind}-{index}",
            "kind": kind,
            "category": normalize_text(record.get("类别")),
            "negotiationType": normalize_text(record.get("谈判药/竞价药")),
            "code": normalize_text(record.get("编码")),
            "name": normalize_text(record.get("名称")),
            "dosageForm": normalize_text(record.get("剂型")),
            "textType": normalize_text(record.get("文本类型")),
            "textContent": normalize_text(record.get("文本内容")),
            "source": normalize_text(record.get("来源")),
        }
        entries.append(entry)
    return entries


def build_service_entries(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    for index, record in enumerate(records, start=1):
        entries.append(
            {
                "id": f"service-{index}",
                "kind": "service",
                "code": normalize_text(record.get("编码")),
                "name": normalize_text(record.get("名称")),
                "textType": normalize_text(record.get("文本类型")),
                "textContent": normalize_text(record.get("文本内容")),
                "source": normalize_text(record.get("来源")),
            }
        )
    return entries


def build_rules(
    definitions: list[dict[str, Any]],
    conditions: list[dict[str, Any]],
    conclusions: list[dict[str, Any]],
    domain: str,
) -> list[dict[str, Any]]:
    conditions_by_rule: dict[str, list[dict[str, Any]]] = {}
    for condition in conditions:
        rule_id = normalize_text(condition.get("ruleId") or condition.get("规则标识符"))
        if not rule_id:
            continue
        conditions_by_rule.setdefault(rule_id, []).append(
            {
                "conditionId": normalize_text(condition.get("conditionId") or condition.get("条件序号")),
                "conditionCombine": normalize_text(condition.get("conditionCombine") or condition.get("条件连接")),
                "groupId": normalize_text(condition.get("groupId") or condition.get("条件组序号")),
                "groupCombine": normalize_text(condition.get("groupCombine") or condition.get("条件组逻辑")),
                "deId": normalize_text(condition.get("deId") or condition.get("数据项标识符")),
                "paramTag": normalize_text(condition.get("paramTag") or condition.get("参数标签")),
                "operator": normalize_text(condition.get("operator") or condition.get("运算符")),
                "value": normalize_text(condition.get("value") or condition.get("判断值")),
                "valueSetId": normalize_text(condition.get("valueSetId") or condition.get("值集标识符")),
                "valueSetName": normalize_text(condition.get("值集名称")),
                "kgProperty": normalize_text(condition.get("kgProperty") or condition.get("知识图谱属性")),
                "ruleName": normalize_text(condition.get("规则名称")),
            }
        )

    conclusions_by_rule: dict[str, list[dict[str, Any]]] = {}
    for conclusion in conclusions:
        rule_id = normalize_text(conclusion.get("ruleId") or conclusion.get("规则ID"))
        if not rule_id:
            continue
        conclusions_by_rule.setdefault(rule_id, []).append(
            {
                "resultType": normalize_text(conclusion.get("resultType") or conclusion.get("结论类型")),
                "resultTip": normalize_text(conclusion.get("resultTip") or conclusion.get("提示内容")),
            }
        )

    rules: list[dict[str, Any]] = []
    for definition in definitions:
        rule_id = normalize_text(definition.get("ruleId") or definition.get("规则标识符"))
        if not rule_id:
            continue
        rules.append(
            {
                "id": rule_id,
                "domain": domain,
                "name": normalize_text(definition.get("ruleName") or definition.get("规则名称")),
                "type": normalize_text(definition.get("ruleType") or definition.get("规则类型")),
                "note": normalize_text(definition.get("ruleNote") or definition.get("规则备注")),
                "splitFlag": normalize_text(definition.get("splitFlag") or definition.get("拆分标记")),
                "originRuleId": normalize_text(definition.get("originRuleId") or definition.get("拆分前规则标识符")),
                "status": normalize_text(definition.get("ruleStatus") or definition.get("规则状态")),
                "source": normalize_text(definition.get("ruleSource") or definition.get("规则来源")),
                "conditions": conditions_by_rule.get(rule_id, []),
                "conclusions": conclusions_by_rule.get(rule_id, []),
            }
        )
    return rules


def main() -> None:
    drug_text_file = next(MATERIALS_DIR.glob("医学规则库_医保用药规则_*/*规则文本*.xlsx"))
    drug_structured_file = next(MATERIALS_DIR.glob("医学规则库_医保用药规则_*/*结构化规则易读版*.xlsx"))
    service_text_file = next(MATERIALS_DIR.glob("医学规则库_医保诊疗项目规则_*/*规则文本*.xlsx"))
    service_structured_file = next(MATERIALS_DIR.glob("医学规则库_医保诊疗项目规则_*/*结构化规则易读版*.xlsx"))

    drug_text_entries = build_text_entries(sheet_rows(drug_text_file, "Sheet1"), "drug")
    service_text_entries = build_service_entries(sheet_rows(service_text_file, "Sheet1"))

    drug_rules = build_rules(
        sheet_rows(drug_structured_file, "规则定义")[2:],
        sheet_rows(drug_structured_file, "规则条件")[2:],
        sheet_rows(drug_structured_file, "规则结论")[2:],
        "drug",
    )
    service_rules = build_rules(
        sheet_rows(service_structured_file, "规则定义")[2:],
        sheet_rows(service_structured_file, "规则条件")[2:],
        sheet_rows(service_structured_file, "规则结论")[2:],
        "service",
    )

    payload = {
        "generatedAt": "2026-04-26",
        "sources": {
            "drugText": str(drug_text_file),
            "drugStructured": str(drug_structured_file),
            "serviceText": str(service_text_file),
            "serviceStructured": str(service_structured_file),
        },
        "drugDirectory": drug_text_entries,
        "serviceDirectory": service_text_entries,
        "drugRules": drug_rules,
        "serviceRules": service_rules,
        "ruleLibrary": drug_rules + service_rules,
    }

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"generated: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
